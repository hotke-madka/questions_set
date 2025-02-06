import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
import os
import re

#0522　オリジナル画像の表示を撤廃を目指す。代わりにentry_sectionにopen_phto_btnを追加する。
class ImageDisplayApp:
    def __init__(self, master):
        self.master = master
        self.master.title("アンケート入力ツール")
        self.master.geometry("1350x500")
        
        self.input_field = []
        self.image_paths = []
        self.original_images = []
        self.question_images = []
        
        self.questions = ["質問1", "質問2", "質問3","質問4","質問5", "質問6", "質問7","質問8","質問9", "質問10", "質問11","質問12","質問13", "質問14", "質問15","質問16","質問17"]
        self.current_question_index = 0
        self.current_paper_index = 1
        self.answers = [[] for _ in range(len(self.questions))] 
        self.excel_file_path = None
        self.current_image_index = 0
        
        self.display_option()
        
        self.select_excel_file()
        self.select_folders()
        self.load_images()
        

    def display_option(self):
        #self.original_image_frame = tk.Frame(self.master)
        #self.original_image_frame.pack(side="left", padx=10, pady=10)
        self.questions_image_frame = tk.Frame(self.master, width=1300, height=300)
        self.questions_image_frame.pack(side="top", padx=10, pady=10)
        self.input_frame = tk.Frame(self.master)
        self.input_frame.pack(side="top", padx=10, pady=10)
        
        self.entry_section(self.input_frame)
        
    def entry_section(self, parent_frame):

        self.question_label = tk.Label(parent_frame, text=self.questions[self.current_question_index])
        self.question_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        self.answer_entry = tk.Entry(parent_frame)
        self.answer_entry.grid(row=0, column=1, padx=10, pady=10)
        
        self.answer_entry.bind("<Return>", self.confirm_answer)
        self.confirm_btn = tk.Button(parent_frame, text="回答確定", command=self.confirm_answer)
        self.confirm_btn.grid(row=1, column=0, padx=10, pady=10)

        self.paper_index_label = tk.Label(parent_frame, text=f"{self.current_paper_index}枚目のアンケート")
        self.paper_index_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        self.photo_btn = tk.Button(parent_frame, text="画像を表示", command=self.open_original)
        self.photo_btn.grid(row=4, column=0, padx=10, pady=10)
        
    def confirm_answer(self,Event=None):
        answer = self.answer_entry.get()
        self.answers[self.current_question_index].append(answer)
        print(f"質問{self.current_question_index+1}への回答({self.current_paper_index}枚目のアンケート): {answer}")
        
        self.current_question_index += 1
        
        if self.current_question_index < len(self.questions):
            self.update_question_entry()
        else:
            self.current_question_index = 0
            self.current_paper_index += 1
            print(f"{self.current_paper_index-1}枚目のアンケートを記録しました")
            self.save_to_excel()
            self.answers = [[] for _ in range(len(self.questions))]
            self.update_question_entry()
            self.paper_index_label.config(text=f"{self.current_paper_index}枚目のアンケート")
            
            if hasattr(self, "original_label"):
                self.original_label.destroy()
            
            self.show_next_original_image()
        
        # 次の質問画像を表示
        self.show_next_question_image()
    
    def open_original(self):
        file_path = self.original_image_paths[self.current_paper_index-1]
        if os.path.exists(file_path):
            os.startfile(file_path)
        else:
            print("指定されたファイルは存在しません。")
            
    def show_next_question_image(self):
        if self.current_image_index <= len(self.question_images):
            self.current_image_index += 1
            current_question_image = self.question_images[self.current_image_index]
            
            if hasattr(self, "question_image_label"):
                self.question_image_label.destroy()
            self.question_image_label = tk.Label(self.questions_image_frame, image=current_question_image)
            self.question_image_label.image = current_question_image
            self.question_image_label.pack()            
        else:
            messagebox.showinfo("情報", "すべてのアンケートを記録しました")
    
    def show_next_original_image(self):
        current_original_image = self.original_images[self.current_paper_index-1]
        
        #if hasattr(self, "question_image_label"):
            #self.original_image_label.destroy()
        #self.original_image_label = tk.Label(self.original_image_frame, image=current_original_image)
        #self.original_image_label.image = current_original_image
        #self.original_image_label.pack()
            
    def update_question_entry(self):
        self.question_label.config(text=self.questions[self.current_question_index])
        self.answer_entry.delete(0, "end")
        self.answer_entry.focus_set()

    def save_to_excel(self):
        try:
            if self.excel_file_path:
                wb = openpyxl.load_workbook(self.excel_file_path)
                ws = wb.active
                
                for paper_answers in zip(*self.answers):
                    ws.append([f"id{self.current_paper_index-1}"] + list(paper_answers))
                wb.save(self.excel_file_path)
                print(f"回答をExcelファイルに追加しました: {self.excel_file_path}")
            else:
                messagebox.showwarning("警告", "Excelファイルが選択されていません。")
        except Exception as e:
            messagebox.showerror("エラー", f"回答をExcelファイルに記録する際にエラーが発生しました: {str(e)}")
            
    def select_excel_file(self):
        self.excel_file_path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if self.excel_file_path:
            self.entry_section(self.input_frame)
            self.load_existing_papers()
    
    def load_existing_papers(self):
        try:
            if self.excel_file_path:
                wb = openpyxl.load_workbook(self.excel_file_path)
                ws = wb.active
                self.current_paper_index = ws.max_row
                self.paper_index_label.config(text=f"{self.current_paper_index}枚目のアンケート") 
                print(f"既存のアンケートをロードしました。{self.current_paper_index}枚目のアンケート")
        except Exception as e:
            messagebox.showerror("エラー", f"既存のアンケートの読み込みに失敗しました: {str(e)}")
            
    def select_folders(self):
        self.original_folder_path = filedialog.askdirectory(title="オリジナル画像フォルダを選択")
        self.question_folder_path = filedialog.askdirectory(title="質問画像フォルダを選択")
        
    def load_images(self):
        if self.original_folder_path and self.question_folder_path:
            self.load_original_images()
            self.load_question_images()
            self.show_images()
        else:
            messagebox.showwarning("警告", "オリジナル画像フォルダと質問画像フォルダの両方を選択してください。")
            
    def load_original_images(self):
        # オリジナル画像フォルダから画像ファイルを取得してソート
        original_image_files = sorted([file for file in os.listdir(self.original_folder_path) if file.endswith((".jpg", ".jpeg", ".png"))],
                                    key=lambda x: int(re.findall(r'id(\d+)', x)[0]))
        
        # ソートされた画像ファイルのパスを生成
        self.original_image_paths = [os.path.join(self.original_folder_path, file) for file in original_image_files]
        # 画像をロード
        self.original_images = [self.load_original_image(path) for path in self.original_image_paths]

        
    def load_question_images(self):
        # フォルダ内の画像ファイルを取得
        question_image_files = sorted([file for file in os.listdir(self.question_folder_path) if file.endswith((".jpg", ".jpeg", ".png"))])
        
        # idとreg番号に基づいて画像をソート
        question_image_files.sort(key=lambda x: (int(re.findall(r'id(\d+)', x)[0]), int(re.findall(r'reg(\d+)', x)[0])))
        
        # ソートされた画像をロード
        self.question_image_paths = [os.path.join(self.question_folder_path, file) for file in question_image_files]
        self.question_images = [self.load_question_image(path) for path in self.question_image_paths]

    
    def load_original_image(self, image_path):
        if os.path.exists(image_path):
            original_image = Image.open(image_path)
            original_image = original_image.resize((300, 300))
            original_photo = ImageTk.PhotoImage(original_image)
            return original_photo
        else:
            return None
    
    def load_question_image(self, image_path):
        if os.path.exists(image_path):
            question_image = Image.open(image_path)
            question_photo = ImageTk.PhotoImage(question_image)
            return question_photo
        else:
            return None
    
    def show_images(self):
        if self.original_images and self.question_images:
            if self.current_paper_index <= len(self.original_images):
                self.current_image_index = (len(self.questions)) * (self.current_paper_index-1)
                current_original_image = self.original_images[self.current_paper_index-1]
                current_question_image = self.question_images[self.current_image_index]
                print(self.current_paper_index)
                print(self.current_image_index)
                
                if hasattr(self, "question_image_label"):
                    self.question_image_label.destroy()
                self.question_image_label = tk.Label(self.questions_image_frame, image=current_question_image)
                self.question_image_label.image = current_question_image
                self.question_image_label.pack()
                
                self.paper_original_label = tk.Label(self.input_frame, text=f"全部で({len(self.original_images)})枚")
                self.paper_original_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")
            
            else:
                messagebox.showinfo("完了","全てのアンケートを記録しました")
            
            #if current_original_image:
             #   if hasattr(self, "original_image_label"):
              #      self.original_image_label.destroy()
                #self.original_image_label = tk.Label(self.original_image_frame, image=current_original_image)
                #self.original_image_label.image = current_original_image
                #self.original_image_label.pack()


def main():
    master = tk.Tk()
    app = ImageDisplayApp(master)
    master.mainloop()
    
if __name__ =="__main__":
    main()
